# -*- coding: utf-8 -*-
# Import necessary libraries
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import numpy as np
import datetime
import random
import os
import configparser
import warnings
import traceback # For detailed error logging
from openpyxl.styles import Font # Import Font for Excel styling

# --- Configuration & Setup ---

# Ignore openpyxl warnings (often related to default styles)
warnings.simplefilter("ignore", category=UserWarning)
# Suppress SettingWithCopyWarning (use .loc/.copy() appropriately)
pd.options.mode.chained_assignment = None

# --- Constants ---
DEFAULT_REVIEWERS = ["Prabhakar", "Sudhakar", "Anthoni", "Rajalakshmi", "Narasimha", "Janakiram"]
CONFIG_FILE = 'reviewers.ini'
CONFIG_SECTION = 'Reviewers'

# Define common column names for consistency and easier changes
COL_AER = 'AER#'
COL_REPORT_CLASS = 'Report Classification'
COL_ASSIGNED_TO = 'Assigned To'
COL_COMPANY_UNIT = 'Company Unit'
COL_INDIVIDUAL_ASSIGNMENT = 'Individual Assignment'
COL_REMARKS = 'Remarks'
COL_DUE_DAYS = 'No of days due to Case Due Date'
COL_CASE_SERIOUSNESS = 'Case seriousness'
COL_REPORT_TYPE = 'Report Type'
COL_CASE_DUE_DATE = 'Case Due Date'

# --- Required Columns for Validation ---
# Columns absolutely required in the Lifesphere Export sheet
REQUIRED_LS_COLS = [
    COL_AER, COL_REPORT_CLASS, COL_ASSIGNED_TO, COL_COMPANY_UNIT, COL_DUE_DAYS
]
# Columns absolutely required in BOTH Previous Assignment sheets for the pending check
REQUIRED_PREV_COLS = [
    COL_AER, COL_INDIVIDUAL_ASSIGNMENT
]
# Additional columns needed ONLY from the Previous Priority sheet for balancing logic
REQUIRED_PREV_PRIO_EXTRA_COLS = [
    COL_REPORT_CLASS
]

# --- Configuration Handling Functions ---
def load_reviewers():
    """Loads reviewers from the config file."""
    config = configparser.ConfigParser()
    if os.path.exists(CONFIG_FILE):
        try:
            config.read(CONFIG_FILE)
            reviewers_str = config.get(CONFIG_SECTION, 'list', fallback=','.join(DEFAULT_REVIEWERS))
            return [r.strip() for r in reviewers_str.split(',') if r.strip()]
        except Exception as e:
            print(f"Error loading config file '{CONFIG_FILE}': {e}")
            return DEFAULT_REVIEWERS[:]
    return DEFAULT_REVIEWERS[:]

def save_reviewers(reviewer_list):
    """Saves the current reviewer list to the config file."""
    config = configparser.ConfigParser()
    config[CONFIG_SECTION] = {'list': ','.join(reviewer_list)}
    try:
        with open(CONFIG_FILE, 'w') as configfile:
            config.write(configfile)
        return True
    except Exception as e:
        messagebox.showerror("Config Error", f"Failed to save reviewers list to '{CONFIG_FILE}': {e}")
        return False

# --- Helper Function for Column Check ---
def check_columns(df_columns, required_columns, file_description, sheet_name):
    """Checks if required columns exist in the DataFrame columns."""
    missing_cols = [col for col in required_columns if col not in df_columns]
    if missing_cols:
        error_msg = (f"Error: Missing required columns in {file_description} "
                     f"(Sheet: '{sheet_name}'):\n-> {', '.join(missing_cols)}\n\n"
                     f"Please ensure the column names match exactly (case-sensitive).")
        return False, error_msg
    return True, "" # Success

# Load initial reviewers
current_reviewers = load_reviewers()

# --- Core Assignment Logic (with enhanced debugging) ---
def perform_assignment(lifesphere_file, prev_assign_file, ls_sheet_name, prev_prio_sheet_name, prev_pend_sheet_name, selected_reviewers, output_dir):
    """
    Performs the case assignment logic based on the input files and parameters.
    Returns: (success_boolean, message_string)
    """
    print("\n--- Starting Assignment Process ---")
    print(f"Lifesphere File: {lifesphere_file} (Sheet: '{ls_sheet_name}')")
    print(f"Previous Assignment File: {prev_assign_file} (Priority Sheet: '{prev_prio_sheet_name}', Pending Sheet: '{prev_pend_sheet_name}')")
    print(f"Selected Reviewers: {selected_reviewers}")
    print(f"Output Directory: {output_dir}")

    # <<<--- USER: REPLACE with 2-3 actual AER#s that are failing --->>>
    # Example: problematic_aers_debug = ['1234567', '9876543']
    problematic_aers_debug = ['AER_ID_1', 'AER_ID_2', 'AER_ID_9'] # Replace with actual AER#s
    print(f"DEBUG: Will specifically track these AER#s: {problematic_aers_debug}")

    try:
        # --- Step 1: Load Lifesphere Data ---
        print("\nStep 1: Loading Lifesphere data...")
        try:
            df_ls_raw = pd.read_excel(lifesphere_file, sheet_name=ls_sheet_name, skiprows=5, dtype={COL_AER: str})
            df_master_data = df_ls_raw.copy()
            print(f"Successfully loaded {len(df_ls_raw)} rows from '{ls_sheet_name}'.")
        except FileNotFoundError:
            return False, f"Error: Lifesphere file not found at {lifesphere_file}"
        except ValueError as e:
             if f"Worksheet named '{ls_sheet_name}' not found" in str(e):
                 return False, f"Error: Worksheet named '{ls_sheet_name}' not found in the Lifesphere file. Please check the sheet name (case-sensitive)."
             else:
                return False, f"Error reading Lifesphere sheet '{ls_sheet_name}': {e}. Check file format/corruption."
        except Exception as e:
            return False, f"An unexpected error occurred reading Lifesphere file: {e}\n{traceback.format_exc()}"

        # --- Verify Required Columns ---
        required_ls_cols = [ COL_AER, COL_REPORT_CLASS, COL_ASSIGNED_TO, COL_COMPANY_UNIT, COL_DUE_DAYS ]
        critical_missing = [col for col in required_ls_cols if col not in df_ls_raw.columns]
        if critical_missing:
             return False, f"Error: Missing CRITICAL columns in Lifesphere sheet ('{ls_sheet_name}'): {', '.join(critical_missing)}"

        df = df_ls_raw.copy()
        # Ensure critical AER column is string and stripped early
        df[COL_AER] = df[COL_AER].astype(str).str.strip().replace('nan', '') # Replace 'nan' string
        df.dropna(subset=[COL_AER], inplace=True)
        df = df[df[COL_AER] != ''] # Remove rows where AER# became empty
        print(f"Cleaned {COL_AER}, rows remaining after dropna/empty removal: {len(df)}")


        # --- FILTER: Company Unit ---
        print(f"\nFiltering by '{COL_COMPANY_UNIT}' (keeping 'BMS')...")
        original_count = len(df)
        df = df[df[COL_COMPANY_UNIT].astype(str).str.contains('BMS', case=False, na=False)].copy()
        print(f"Rows after '{COL_COMPANY_UNIT}' filter: {len(df)} (removed {original_count - len(df)})")


        # --- Step 2: Clean 'Days Due' ---
        print("\nStep 2: Cleaning 'Days Due' column...")
        if COL_DUE_DAYS in df.columns:
            df[COL_DUE_DAYS] = df[COL_DUE_DAYS].astype(str).str.replace(r'\s*day\(s\)', '', regex=True).str.strip()
            df[COL_DUE_DAYS] = pd.to_numeric(df[COL_DUE_DAYS], errors='coerce')
            print("Cleaned 'Days Due' column.")
        else:
             print(f"Warning: Column '{COL_DUE_DAYS}' not found. Sorting by this column will be skipped.")


        # --- Step 3: Filter 'Report Classification' ---
        print(f"\nStep 3: Filtering by '{COL_REPORT_CLASS}' (Literature, Non-AE Case, Blank)...")
        original_count = len(df)
        filter_values = ["Literature", "Non-AE Case"]
        df = df[
            df[COL_REPORT_CLASS].isin(filter_values) |
            df[COL_REPORT_CLASS].isnull() |
            (df[COL_REPORT_CLASS].astype(str).str.strip() == '')
        ].copy()
        print(f"Rows after '{COL_REPORT_CLASS}' filter: {len(df)} (removed {original_count - len(df)})")


        # --- Step 4: Filter 'Assigned To' ---
        print(f"\nStep 4: Filtering by '{COL_ASSIGNED_TO}' (removing emails)...")
        original_count = len(df)
        df[COL_ASSIGNED_TO] = df[COL_ASSIGNED_TO].fillna('')
        df = df[
            (df[COL_ASSIGNED_TO] == '') |
            (~df[COL_ASSIGNED_TO].astype(str).str.contains('@', na=False))
        ].copy()
        print(f"Rows after '{COL_ASSIGNED_TO}' filter: {len(df)} (removed {original_count - len(df)})")


        # --- Step 5: Prepare master data frame ---
        print("\nStep 5: Preparing master data frame for processing...")
        df_priority_master = df.copy()
        # Ensure AER# is clean just before merge step as well
        df_priority_master[COL_AER] = df_priority_master[COL_AER].astype(str).str.strip().replace('nan', '')
        df_priority_master = df_priority_master[df_priority_master[COL_AER] != ''] # Ensure no empty AERs here
        df_priority_master[COL_INDIVIDUAL_ASSIGNMENT] = pd.NA
        df_priority_master[COL_REMARKS] = pd.NA
        print(f"Initial 'df_priority_master' created with {len(df_priority_master)} rows.")


        # --- Step 6: Load and Combine Previous Assignments (Enhanced Debugging) ---
        print("\nStep 6: Loading and combining previous assignment data...")
        df_pending_caselist = pd.DataFrame() # Combined list
        df_prev_priority_raw = pd.DataFrame() # Raw prev priority for step 9

        if prev_assign_file and os.path.exists(prev_assign_file):
            print(f"Processing previous assignment file: {os.path.basename(prev_assign_file)}")
            # Load sheets with explicit string type for AER#
            try:
                df_prev_priority = pd.read_excel(prev_assign_file, sheet_name=prev_prio_sheet_name, dtype={COL_AER: str})
                df_prev_priority_raw = df_prev_priority.copy()
                print(f"  Read {len(df_prev_priority)} rows from '{prev_prio_sheet_name}'.")
            except ValueError as e:
                 print(f"  ERROR reading '{prev_prio_sheet_name}': {e}")
                 df_prev_priority = pd.DataFrame()
            except Exception as e:
                print(f"  ERROR reading '{prev_prio_sheet_name}': {e}")
                df_prev_priority = pd.DataFrame()

            try:
                df_prev_pending = pd.read_excel(prev_assign_file, sheet_name=prev_pend_sheet_name, dtype={COL_AER: str})
                print(f"  Read {len(df_prev_pending)} rows from '{prev_pend_sheet_name}'.")
            except ValueError as e:
                 print(f"  ERROR reading '{prev_pend_sheet_name}': {e}")
                 df_prev_pending = pd.DataFrame()
            except Exception as e:
                print(f"  ERROR reading '{prev_pend_sheet_name}': {e}")
                df_prev_pending = pd.DataFrame()

            # Process and Combine
            req_prev_cols = [COL_AER, COL_INDIVIDUAL_ASSIGNMENT]
            processed_dfs = []

            def process_prev_df(df, sheet_name):
                # Process previous dataframes rigorously
                if df is None or df.empty:
                    print(f"  Skipping processing for empty DataFrame ({sheet_name}).")
                    return pd.DataFrame()
                print(f"  Processing DataFrame from '{sheet_name}'...")
                required_cols_here = [COL_AER, COL_INDIVIDUAL_ASSIGNMENT] # Need assignment to eventually populate if pending
                missing = [col for col in required_cols_here if col not in df.columns]
                if missing:
                    print(f"  WARNING: Missing columns in '{sheet_name}': {', '.join(missing)}. Skipping.")
                    return pd.DataFrame()

                df_proc = df[required_cols_here].copy()

                # Rigorous Cleaning of AER# - Essential for matching
                df_proc[COL_AER] = df_proc[COL_AER].astype(str).str.strip().replace('nan', '')
                # Drop rows ONLY if AER# became empty/invalid after cleaning
                df_proc.dropna(subset=[COL_AER], inplace=True) # Drop if actual NaN
                df_proc = df_proc[df_proc[COL_AER] != '']      # Drop if empty string

                # Clean Assignment column but DON'T drop row if it's empty, just fill later if needed
                df_proc[COL_INDIVIDUAL_ASSIGNMENT] = df_proc[COL_INDIVIDUAL_ASSIGNMENT].fillna('').astype(str).str.strip().replace('nan', '')
                # Fill empty assignments with a placeholder IF NEEDED for merge logic (though merge picks first value anyway)
                # df_proc.loc[df_proc[COL_INDIVIDUAL_ASSIGNMENT] == '', COL_INDIVIDUAL_ASSIGNMENT] = 'Unknown_Prev_Assign' # Optional

                if not df_proc.empty:
                    print(f"  Valid rows after cleaning AER# from '{sheet_name}': {len(df_proc)}")
                    # DEBUG Check
                    found_debug = df_proc[df_proc[COL_AER].isin(problematic_aers_debug)]
                    if not found_debug.empty: print(f"  DEBUG: Found problematic AERs in processed '{sheet_name}':\n{found_debug}")
                    return df_proc
                else:
                    print(f"  No valid rows remaining after cleaning AER# from '{sheet_name}'.")
                    return pd.DataFrame()

            df_prio_processed = process_prev_df(df_prev_priority, prev_prio_sheet_name)
            df_pend_processed = process_prev_df(df_prev_pending, prev_pend_sheet_name)

            # Combine (Priority cases added first)
            if not df_prio_processed.empty: processed_dfs.append(df_prio_processed)
            if not df_pend_processed.empty: processed_dfs.append(df_pend_processed)

            if processed_dfs:
                df_pending_caselist = pd.concat(processed_dfs, ignore_index=True)
                print(f"  Combined previous assignments before deduplication: {len(df_pending_caselist)} rows.")
                # Keep the first entry for any duplicate AER# (Priority takes precedence)
                df_pending_caselist = df_pending_caselist.drop_duplicates(subset=[COL_AER], keep='first')
                print(f"  Final unique previous assignment list size: {len(df_pending_caselist)} rows.")
                found_final_debug = df_pending_caselist[df_pending_caselist[COL_AER].isin(problematic_aers_debug)]
                if not found_final_debug.empty: print(f"  DEBUG: Problematic AERs in FINAL combined previous list:\n{found_final_debug}")
                else: print(f"  DEBUG: Problematic AERs {problematic_aers_debug} NOT FOUND in final combined previous list.")
            else:
                print("  No valid data found in previous assignment sheets to combine.")

            if COL_REPORT_CLASS not in df_prev_priority_raw.columns:
                 print(f"  WARNING: Column '{COL_REPORT_CLASS}' missing in '{prev_prio_sheet_name}'. Balancing (Step 9) may be affected.")

        else:
             print("Previous assignment file not provided or not found.")


        # --- Step 7 & 8 Combined: Identify previously assigned and mark (Enhanced Debugging) ---
        # --- Step 7 & 8 Combined: Identify previously assigned and mark (Enhanced Debugging) ---
        print("\nSteps 7 & 8: Identifying pending cases using combined previous list...")
        if not df_pending_caselist.empty:
            # --- PRE-MERGE CHECKS ---
            print("  --- Pre-Merge Debug Info ---")
            print(f"  Master DF rows: {len(df_priority_master)}, Prev List rows: {len(df_pending_caselist)}")
            print(f"  Master AER# dtype: {df_priority_master[COL_AER].dtype}, Prev List AER# dtype: {df_pending_caselist[COL_AER].dtype}")
            # Ensure keys are clean just before merge
            df_priority_master[COL_AER] = df_priority_master[COL_AER].astype(str).str.strip()
            df_pending_caselist[COL_AER] = df_pending_caselist[COL_AER].astype(str).str.strip()

            master_debug_pre = df_priority_master[df_priority_master[COL_AER].isin(problematic_aers_debug)][[COL_AER, COL_INDIVIDUAL_ASSIGNMENT, COL_REMARKS]]
            if not master_debug_pre.empty: print(f"  DEBUG: Problematic AERs in Master DF BEFORE merge:\n{master_debug_pre}")
            else: print(f"  DEBUG: Problematic AERs {problematic_aers_debug} NOT FOUND in Master DF before merge.")

            prev_debug_pre = df_pending_caselist[df_pending_caselist[COL_AER].isin(problematic_aers_debug)]
            if not prev_debug_pre.empty: print(f"  DEBUG: Problematic AERs in Prev List BEFORE merge:\n{prev_debug_pre}")
            else: print(f"  DEBUG: Problematic AERs {problematic_aers_debug} NOT FOUND in Prev List before merge.")

            # --- PERFORM MERGE ---
            print("  Performing merge...")
            try:
                # Assign merge result to a new variable to avoid modifying df_priority_master inplace yet
                df_merged_temp = pd.merge(
                    df_priority_master,
                    df_pending_caselist[[COL_AER, COL_INDIVIDUAL_ASSIGNMENT]], # Select only needed columns
                    on=COL_AER,
                    how='left',
                    suffixes=('', '_prev'), # Suffix important
                    indicator=True # Crucial for checking merge success
                )
                print("  Merge completed.")
                print(f"  Merge result value counts:\n{df_merged_temp['_merge'].value_counts()}")
            except Exception as merge_err:
                print(f"  ERROR during merge: {merge_err}")
                return False, f"Critical error during merge operation: {merge_err}"

            # --- POST-MERGE CHECKS ---
            print("  --- Post-Merge Debug Info ---")
            merged_debug_post = df_merged_temp[df_merged_temp[COL_AER].isin(problematic_aers_debug)][[COL_AER, COL_INDIVIDUAL_ASSIGNMENT, COL_INDIVIDUAL_ASSIGNMENT + '_prev', COL_REMARKS, '_merge']]
            if not merged_debug_post.empty: print(f"  DEBUG: Problematic AERs in Merged DF:\n{merged_debug_post}")
            else: print(f"  DEBUG: Problematic AERs {problematic_aers_debug} NOT FOUND in Merged DF.")

            # Identify rows where a match was found (_merge == 'both')
            previously_assigned_mask = df_merged_temp['_merge'] == 'both'
            num_found_pending = previously_assigned_mask.sum()
            print(f"  Identified {num_found_pending} cases as potentially pending based on merge indicator ('both').")

            mask_debug = df_merged_temp.loc[df_merged_temp[COL_AER].isin(problematic_aers_debug), '_merge']
            if not mask_debug.empty: print(f"  DEBUG: Merge status for problematic AERs:\n{mask_debug}")

            # --- UPDATE BASED ON MASK ---
            # Create copies to avoid SettingWithCopyWarning when updating based on conditions
            df_priority_master_updated = df_merged_temp.copy()

            if num_found_pending > 0:
                print(f"  Updating {num_found_pending} rows based on merge matches...")
                # Apply updates directly to the copied merged dataframe
                df_priority_master_updated.loc[previously_assigned_mask, COL_INDIVIDUAL_ASSIGNMENT] = df_priority_master_updated.loc[previously_assigned_mask, COL_INDIVIDUAL_ASSIGNMENT + '_prev']
                df_priority_master_updated.loc[previously_assigned_mask, COL_REMARKS] = 'Pending from prev. allocation'
                print(f"  Marked {num_found_pending} cases as 'Pending from prev. allocation' in the temporary merged df.")

                # --- POST-UPDATE CHECK ---
                updated_debug_post = df_priority_master_updated.loc[df_priority_master_updated[COL_AER].isin(problematic_aers_debug), [COL_AER, COL_INDIVIDUAL_ASSIGNMENT, COL_REMARKS, '_merge']]
                if not updated_debug_post.empty: print(f"  DEBUG: Problematic AERs in updated temp DF:\n{updated_debug_post}")
                else: print(f"  DEBUG: Problematic AERs {problematic_aers_debug} NOT FOUND after update (unexpected?).")

            else:
                print("  No cases found with merge status 'both', no updates applied.")

            # Clean up temporary columns from the updated df
            df_priority_master_updated.drop(columns=['_merge', COL_INDIVIDUAL_ASSIGNMENT + '_prev'], inplace=True, errors='ignore')

            # Assign the fully processed DataFrame back
            df_priority_master = df_priority_master_updated
            print("  Assigned updated merged data back to df_priority_master.")

        else:
            print("  No previous assignment data to perform matching.")


        # --- Filter for cases needing assignment THIS round ---
        # Use the FINAL updated df_priority_master
        df_to_assign = df_priority_master[df_priority_master[COL_INDIVIDUAL_ASSIGNMENT].isna()].copy()
        print(f"\nCases remaining for new assignment in this run: {len(df_to_assign)}")
        # DEBUG: Check if problematic AERs are wrongly in df_to_assign
        to_assign_debug = df_to_assign[df_to_assign[COL_AER].isin(problematic_aers_debug)][[COL_AER, COL_INDIVIDUAL_ASSIGNMENT, COL_REMARKS]]
        if not to_assign_debug.empty: print(f"  DEBUG: *** PROBLEM *** Problematic AERs found in df_to_assign (should be pending):\n{to_assign_debug}")
        else: print(f"  DEBUG: Problematic AERs correctly excluded from df_to_assign.")

        # ... rest of the function ...


        # --- Assignment Steps 9 & 10 ---
        if not selected_reviewers: return False, "Error: No reviewers selected for assignment."
        if df_to_assign.empty: print("\nInfo: No new cases found requiring assignment in this run.")
        else:
            n_reviewers = len(selected_reviewers)
            print(f"\nAssigning {len(df_to_assign)} new cases to {n_reviewers} reviewers: {selected_reviewers}")
            # Step 9: Assign Literature Cases
            print("\nStep 9: Assigning 'Literature' cases...")
            df_lit = df_to_assign[df_to_assign[COL_REPORT_CLASS] == 'Literature'].copy()
            if not df_lit.empty:
                # ... (Balancing and Assignment Logic for Literature - unchanged) ...
                indices_lit = df_lit.index
                num_lit = len(df_lit)
                print(f"  Found {num_lit} Literature cases to assign.")
                prev_lit_assignments = {}
                if not df_prev_priority_raw.empty and COL_INDIVIDUAL_ASSIGNMENT in df_prev_priority_raw.columns and COL_REPORT_CLASS in df_prev_priority_raw.columns:
                     prev_lit_df = df_prev_priority_raw[ (df_prev_priority_raw[COL_REPORT_CLASS] == 'Literature') & (df_prev_priority_raw[COL_INDIVIDUAL_ASSIGNMENT].isin(selected_reviewers)) ].copy()
                     if not prev_lit_df.empty:
                        prev_lit_counts = prev_lit_df[COL_INDIVIDUAL_ASSIGNMENT].value_counts()
                        for reviewer in selected_reviewers: prev_lit_assignments[reviewer] = prev_lit_counts.get(reviewer, 0)
                        print(f"  Previous Literature counts for balancing: {prev_lit_assignments}")
                     else: print("  No previous 'Literature' cases found assigned to selected reviewers for balancing.")

                assignments_lit = {reviewer: 0 for reviewer in selected_reviewers}
                if prev_lit_assignments: # Balancing logic
                    max_prev_count = max(prev_lit_assignments.values()) if prev_lit_assignments else 0
                    base_lit, remainder_lit = divmod(num_lit, n_reviewers)
                    for r in selected_reviewers: assignments_lit[r] = base_lit
                    reviewers_at_max = [r for r, count in prev_lit_assignments.items() if count == max_prev_count]
                    reviewers_below_max = [r for r in selected_reviewers if r not in reviewers_at_max]
                    random.shuffle(reviewers_below_max)
                    eligible_for_remainder = reviewers_below_max + reviewers_at_max
                    for i in range(remainder_lit):
                        if not eligible_for_remainder: break
                        reviewer_to_get_extra = eligible_for_remainder[i % len(eligible_for_remainder)]
                        assignments_lit[reviewer_to_get_extra] += 1
                else: # Equal assignment
                    base_lit, remainder_lit = divmod(num_lit, n_reviewers)
                    shuffled_reviewers = selected_reviewers[:]
                    random.shuffle(shuffled_reviewers)
                    for i, reviewer in enumerate(shuffled_reviewers): assignments_lit[reviewer] = base_lit + (1 if i < remainder_lit else 0)

                lit_indices_shuffled = list(indices_lit)
                random.shuffle(lit_indices_shuffled)
                current_idx_pos = 0
                for reviewer, count in assignments_lit.items():
                    if current_idx_pos >= len(lit_indices_shuffled): break
                    assign_indices = lit_indices_shuffled[current_idx_pos : current_idx_pos + count]
                    df_priority_master.loc[assign_indices, COL_INDIVIDUAL_ASSIGNMENT] = reviewer # Assign back to main DF
                    current_idx_pos += count
                print(f"  Applied Literature assignment counts: {assignments_lit}")
            else: print("  No new Literature cases to assign.")

            # Step 10: Assign Blanks & Non-AE Cases
            print("\nStep 10: Assigning 'Non-AE Case' & 'Blank' cases...")
            df_other = df_to_assign[ (df_to_assign[COL_REPORT_CLASS] == 'Non-AE Case') | (df_to_assign[COL_REPORT_CLASS].isnull()) | (df_to_assign[COL_REPORT_CLASS].astype(str).str.strip() == '') ].copy()
            if not df_other.empty:
                 # ... (Assignment Logic for Blank/Non-AE Case- unchanged) ...
                 indices_other = df_other.index
                 num_other = len(df_other)
                 print(f"  Found {num_other} Non-AE Case/Blank cases to assign.")
                 base_other, remainder_other = divmod(num_other, n_reviewers)
                 assignments_other = {reviewer: 0 for reviewer in selected_reviewers}
                 shuffled_reviewers = selected_reviewers[:]
                 random.shuffle(shuffled_reviewers)
                 for i, reviewer in enumerate(shuffled_reviewers): assignments_other[reviewer] = base_other + (1 if i < remainder_other else 0)

                 other_indices_shuffled = list(indices_other)
                 random.shuffle(other_indices_shuffled)
                 current_idx_pos = 0
                 for reviewer, count in assignments_other.items():
                     if current_idx_pos >= len(other_indices_shuffled): break
                     assign_indices = other_indices_shuffled[current_idx_pos : current_idx_pos + count]
                     df_priority_master.loc[assign_indices, COL_INDIVIDUAL_ASSIGNMENT] = reviewer # Assign back to main DF
                     current_idx_pos += count
                 print(f"  Applied Blank/Non-AE Case assignment counts: {assignments_other}")
            else: print("  No new Blank/Non-AE Case cases to assign.")


        # --- Step 11: Prepare Output ---
        print("\nStep 11: Preparing output file...")
        today_date_str = datetime.datetime.now().strftime("%d %b %Y")
        output_filename = os.path.join(output_dir, f"{today_date_str}_Assignment.xlsx")

        # Define final output DataFrames using the final df_priority_master
        df_output_master = df_master_data

        df_output_priority = df_priority_master[ df_priority_master[COL_INDIVIDUAL_ASSIGNMENT].notna() & df_priority_master[COL_REMARKS].isna() ].copy()

        # Sort Priority Cases
        if not df_output_priority.empty:
             print("  Sorting Priority Cases output sheet...")
             sort_columns, sort_ascending = [], []
             if COL_REPORT_CLASS in df_output_priority.columns: sort_columns.append(COL_REPORT_CLASS); sort_ascending.append(True)
             if COL_DUE_DAYS in df_output_priority.columns: sort_columns.append(COL_DUE_DAYS); sort_ascending.append(True)
             if sort_columns:
                 df_output_priority = df_output_priority.sort_values(by=sort_columns, ascending=sort_ascending, na_position='last')
                 print(f"  Priority Cases sheet sorted by: {sort_columns}.")
             else: print("  Skipping sorting as required columns are missing.")
        print(f"  Generated 'Priority Cases' sheet with {len(df_output_priority)} rows.")

        df_output_pending = df_priority_master[ df_priority_master[COL_REMARKS] == 'Pending from prev. allocation' ].copy()
        print(f"  Generated 'Pending Cases' sheet with {len(df_output_pending)} rows.")
        # DEBUG: Check if problematic AERs ended up correctly in Pending
        print("  --- Final Output Sheet Debug ---")
        pending_debug_final = df_output_pending[df_output_pending[COL_AER].isin(problematic_aers_debug)][[COL_AER, COL_INDIVIDUAL_ASSIGNMENT, COL_REMARKS]]
        if not pending_debug_final.empty: print(f"  DEBUG: Problematic AERs correctly found in FINAL Pending Cases sheet:\n{pending_debug_final}")
        priority_debug_final = df_output_priority[df_output_priority[COL_AER].isin(problematic_aers_debug)][[COL_AER, COL_INDIVIDUAL_ASSIGNMENT, COL_REMARKS]]
        if not priority_debug_final.empty: print(f"  DEBUG: *** PROBLEM *** Problematic AERs wrongly found in FINAL Priority Cases sheet:\n{priority_debug_final}")


        # Create Summary String
        summary = f"Assignment Summary ({today_date_str}):\n"
        summary += f"- Total cases processed from Lifesphere export ('{ls_sheet_name}'): {len(df_master_data)}\n"
        summary += f"- Cases after initial filters (BMS Unit, Lit/NonAE/Blank, AssignedTo not email): {len(df)}\n"
        summary += f"- Cases identified as pending from previous allocation: {len(df_output_pending)}\n"
        summary += f"- New cases assigned in this run: {len(df_output_priority)}\n"

        # Create Dashboard Pivot Tables
        print("  Generating dashboard data...")
        pivot_priority_user, pivot_priority_class = pd.DataFrame(), pd.DataFrame()
        pivot_pending_user, pivot_pending_class = pd.DataFrame(), pd.DataFrame()

        def create_pivots(df, prefix):
            # ... (create_pivots helper function remains the same) ...
            user_pivot = pd.DataFrame()
            class_pivot = pd.DataFrame()
            if not df.empty:
                try:
                    user_pivot = pd.pivot_table(df, index=COL_INDIVIDUAL_ASSIGNMENT, values=COL_AER, aggfunc='count', fill_value=0)
                    user_pivot.rename(columns={COL_AER: 'Number of Cases'}, inplace=True)
                    # print(f"    {prefix} - User Summary Pivot created.") # Less verbose logging
                except Exception as e: print(f"    WARNING: Failed to create {prefix} User Summary Pivot: {e}")
                try:
                    df['Report Classification Display'] = df[COL_REPORT_CLASS].fillna('Blank').replace('', 'Blank')
                    class_pivot = pd.pivot_table(df, index=['Report Classification Display', COL_INDIVIDUAL_ASSIGNMENT], values=COL_AER, aggfunc='count', fill_value=0)
                    class_pivot.rename(columns={COL_AER: 'Number of Cases'}, inplace=True)
                    # print(f"    {prefix} - Classification Summary Pivot created.")
                except Exception as e: print(f"    WARNING: Failed to create {prefix} Classification Summary Pivot: {e}")
                finally:
                    if 'Report Classification Display' in df.columns:
                        try: df.drop(columns=['Report Classification Display'], inplace=True)
                        except Exception: pass
            # else: print(f"    {prefix} - DataFrame is empty, skipping pivot creation.")
            return user_pivot, class_pivot

        pivot_priority_user, pivot_priority_class = create_pivots(df_output_priority, "Priority Cases")
        if not pivot_priority_user.empty: summary += f"\n\n--- Cases Assigned This Run ({len(df_output_priority)}) ---\nBy Reviewer:\n{pivot_priority_user.to_string()}\n"
        if not pivot_priority_class.empty: summary += f"\nBy Classification and Reviewer:\n{pivot_priority_class.to_string()}\n"
        pivot_pending_user, pivot_pending_class = create_pivots(df_output_pending, "Pending Cases")
        if not pivot_pending_user.empty: summary += f"\n\n--- Cases Pending From Previous ({len(df_output_pending)}) ---\nBy Reviewer:\n{pivot_pending_user.to_string()}\n"
        if not pivot_pending_class.empty: summary += f"\nBy Classification and Reviewer:\n{pivot_pending_class.to_string()}\n"


        # Write to Excel
        print(f"  Writing output file: {output_filename} ...")
        try:
            with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                 # ... (Write data sheets: Master, Priority (sorted), Pending) ...
                 df_output_master.to_excel(writer, sheet_name=f'{today_date_str}_Master data', index=False)
                 df_output_priority.to_excel(writer, sheet_name='Priority Cases', index=False)
                 df_output_pending.to_excel(writer, sheet_name='Pending Cases', index=False)

                 # ... (Write dashboard sheet with bold titles using openpyxl Font) ...
                 dashboard_sheet = writer.book.create_sheet('Dashboard')
                 writer.sheets['Dashboard'] = dashboard_sheet
                 bold_font = Font(bold=True)
                 current_row = 1
                 # Write Priority Pivots
                 if not pivot_priority_user.empty or not pivot_priority_class.empty:
                    title_cell_prio = dashboard_sheet.cell(row=current_row, column=2, value=f"Summary: Cases Assigned This Run ({len(df_output_priority)})"); title_cell_prio.font = bold_font; current_row += 2
                    if not pivot_priority_user.empty: dashboard_sheet.cell(row=current_row, column=2, value="By Reviewer:"); current_row += 1; pivot_priority_user.to_excel(writer, sheet_name='Dashboard', startrow=current_row-1, startcol=1); current_row += pivot_priority_user.shape[0] + 2
                    if not pivot_priority_class.empty: dashboard_sheet.cell(row=current_row, column=2, value="By Classification and Reviewer:"); current_row += 1; pivot_priority_class.to_excel(writer, sheet_name='Dashboard', startrow=current_row-1, startcol=1); current_row += pivot_priority_class.shape[0] + 3
                 # Write Pending Pivots
                 if not pivot_pending_user.empty or not pivot_pending_class.empty:
                    title_cell_pend = dashboard_sheet.cell(row=current_row, column=2, value=f"Summary: Cases Pending From Previous ({len(df_output_pending)})"); title_cell_pend.font = bold_font; current_row += 2
                    if not pivot_pending_user.empty: dashboard_sheet.cell(row=current_row, column=2, value="By Reviewer:"); current_row += 1; pivot_pending_user.to_excel(writer, sheet_name='Dashboard', startrow=current_row-1, startcol=1); current_row += pivot_pending_user.shape[0] + 2
                    if not pivot_pending_class.empty: dashboard_sheet.cell(row=current_row, column=2, value="By Classification and Reviewer:"); current_row += 1; pivot_pending_class.to_excel(writer, sheet_name='Dashboard', startrow=current_row-1, startcol=1); current_row += pivot_pending_class.shape[0] + 2
                 print("  Dashboard sheet generated.")

            print("--- Assignment Process Completed Successfully ---")
            summary += f"\n\nOutput file saved successfully:\n{output_filename}"
            return True, summary
        # ... (Error handling for writing remains the same) ...
        except PermissionError: error_msg = f"Error writing output Excel file: Permission denied. Is '{os.path.basename(output_filename)}' open? Close it and try again."; print(f"  ERROR: {error_msg}"); return False, error_msg
        except Exception as e: error_msg = f"Error writing output Excel file: {e}\n{traceback.format_exc()}"; print(f"  ERROR: {error_msg}"); return False, error_msg

    except Exception as e:
        error_msg = f"An unexpected error occurred during processing: {e}\n{traceback.format_exc()}"
        print(f"--- Assignment Process Failed ---"); print(f"ERROR: {error_msg}")
        return False, error_msg


# --- Tkinter GUI Application Class (Full Version) ---
class CaseAssignerApp:
    def __init__(self, master):
        self.master = master
        master.title("BMS Case Assigner")
        master.geometry("800x750") # Keep adjusted size

        # --- Variables ---
        self.lifesphere_file = tk.StringVar()
        self.prev_assign_file = tk.StringVar()
        self.ls_sheet_name = tk.StringVar(value="Adverse Event") # Default, user MUST verify
        self.prev_prio_sheet_name = tk.StringVar(value="Priority Cases") # Default
        self.prev_pend_sheet_name = tk.StringVar(value="Pending Cases")  # Default
        self.reviewer_vars = {}
        self.all_reviewers = current_reviewers[:]

        # --- Style ---
        style = ttk.Style()
        style.configure('TNotebook.Tab', padding=[10, 5])
        try:
            style.configure('Accent.TButton', font=('Segoe UI', 10, 'bold'), foreground='white', background='#0078D7')
        except tk.TclError:
            print("Warning: Could not configure Accent.TButton style.")


        # --- Notebook for Tabs ---
        self.notebook = ttk.Notebook(master)
        self.main_frame = ttk.Frame(self.notebook, padding="10")
        self.admin_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(self.main_frame, text=' Assignment ')
        self.notebook.add(self.admin_frame, text=' Admin ')
        self.notebook.pack(expand=True, fill="both", padx=5, pady=5)

        # --- Populate Frames ---
        self.create_main_widgets()
        self.create_admin_widgets()

        # --- Summary & Status Area ---
        status_summary_frame = ttk.LabelFrame(master, text="Status & Summary", padding="5")
        status_summary_frame.pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        self.status_label = ttk.Label(status_summary_frame, text="Status: Ready", anchor="w")
        self.status_label.pack(side=tk.TOP, fill=tk.X, padx=5, pady=(0, 5))
        self.summary_text = scrolledtext.ScrolledText(status_summary_frame, height=15, wrap=tk.WORD, state=tk.DISABLED, relief=tk.SUNKEN, bd=1)
        self.summary_text.pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True, padx=5, pady=(0, 5))
        self.show_summary("Welcome! Please select input files and reviewers.\nSheet names are critical - ensure they match your Excel files exactly (case-sensitive).")


    def create_main_widgets(self):
        frame = self.main_frame
        input_frame = ttk.LabelFrame(frame, text="Input Files & Sheets", padding="10")
        input_frame.grid(row=0, column=0, columnspan=2, padx=5, pady=5, sticky="ew")

        ttk.Label(input_frame, text="Lifesphere Export (.xlsx):").grid(row=0, column=0, padx=5, pady=2, sticky="w")
        ttk.Entry(input_frame, textvariable=self.lifesphere_file, width=50).grid(row=0, column=1, padx=5, pady=2, sticky="ew")
        ttk.Button(input_frame, text="Browse...", command=self.browse_lifesphere).grid(row=0, column=2, padx=5, pady=2)
        ttk.Label(input_frame, text="Sheet Name:").grid(row=1, column=0, padx=5, pady=2, sticky="w")
        ttk.Entry(input_frame, textvariable=self.ls_sheet_name, width=25).grid(row=1, column=1, padx=5, pady=2, sticky="w")

        ttk.Label(input_frame, text="Previous Assignment (.xlsx):").grid(row=2, column=0, padx=5, pady=2, sticky="w")
        ttk.Entry(input_frame, textvariable=self.prev_assign_file, width=50).grid(row=2, column=1, padx=5, pady=2, sticky="ew")
        ttk.Button(input_frame, text="Browse...", command=self.browse_previous).grid(row=2, column=2, padx=5, pady=2)
        ttk.Label(input_frame, text="Priority Sheet:").grid(row=3, column=0, padx=5, pady=2, sticky="w")
        ttk.Entry(input_frame, textvariable=self.prev_prio_sheet_name, width=25).grid(row=3, column=1, padx=5, pady=2, sticky="w")
        ttk.Label(input_frame, text="Pending Sheet:").grid(row=4, column=0, padx=5, pady=2, sticky="w")
        ttk.Entry(input_frame, textvariable=self.prev_pend_sheet_name, width=25).grid(row=4, column=1, padx=5, pady=2, sticky="w")

        input_frame.columnconfigure(1, weight=1)

        self.reviewer_frame = ttk.LabelFrame(frame, text="Select Reviewers for Assignment", padding="10")
        self.reviewer_frame.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")
        self.update_reviewer_checkboxes()

        try:
            run_button = ttk.Button(frame, text="Run Assignment", command=self.run_assignment_process)
        except tk.TclError:
            print("Warning: Accent.TButton style not available, using default.")
            run_button = ttk.Button(frame, text="Run Assignment", command=self.run_assignment_process)
        run_button.grid(row=1, column=1, padx=10, pady=10, sticky="se")

        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(1, weight=1)


    def update_reviewer_checkboxes(self):
        for widget in self.reviewer_frame.winfo_children(): widget.destroy()
        self.reviewer_vars.clear()
        self.reviewer_vars = {name: tk.BooleanVar(value=True) for name in self.all_reviewers}
        row, col, max_cols = 0, 0, 3
        for name, var in sorted(self.reviewer_vars.items()):
            cb = ttk.Checkbutton(self.reviewer_frame, text=name, variable=var)
            cb.grid(row=row, column=col, padx=5, pady=2, sticky="w")
            col = (col + 1) % max_cols
            if col == 0: row += 1


    def create_admin_widgets(self):
        frame = self.admin_frame
        list_frame = ttk.LabelFrame(frame, text="Manage Reviewers", padding="5")
        list_frame.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

        self.reviewer_listbox = tk.Listbox(list_frame, height=15)
        self.reviewer_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.reviewer_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.reviewer_listbox.config(yscrollcommand=scrollbar.set)
        self.refresh_reviewer_listbox()

        controls_frame = ttk.Frame(frame, padding="5")
        controls_frame.grid(row=0, column=1, padx=5, pady=5, sticky="ns")

        ttk.Label(controls_frame, text="New Reviewer Name:").pack(pady=(0, 2), anchor="w")
        self.new_reviewer_entry = ttk.Entry(controls_frame, width=25)
        self.new_reviewer_entry.pack(pady=(0, 5), fill=tk.X)
        ttk.Button(controls_frame, text="Add Reviewer", command=self.add_reviewer).pack(fill=tk.X, pady=2)
        ttk.Button(controls_frame, text="Delete Selected", command=self.delete_reviewer).pack(fill=tk.X, pady=2)

        try:
            save_button = ttk.Button(controls_frame, text="Save List", command=self.save_reviewer_list)
        except tk.TclError:
            print("Warning: Accent.TButton style not available, using default for Save button.")
            save_button = ttk.Button(controls_frame, text="Save List", command=self.save_reviewer_list)
        save_button.pack(fill=tk.X, pady=(10, 2))

        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)


    def refresh_reviewer_listbox(self):
        self.reviewer_listbox.delete(0, tk.END)
        for reviewer in sorted(self.all_reviewers):
             self.reviewer_listbox.insert(tk.END, reviewer)


    def add_reviewer(self):
        new_name = self.new_reviewer_entry.get().strip()
        if new_name:
            if new_name not in self.all_reviewers:
                self.all_reviewers.append(new_name)
                self.all_reviewers.sort()
                self.refresh_reviewer_listbox()
                self.update_reviewer_checkboxes()
                self.new_reviewer_entry.delete(0, tk.END)
                self.update_status(f"Admin: Added '{new_name}'. Click 'Save List' to make permanent.")
            else:
                messagebox.showwarning("Duplicate", f"Reviewer '{new_name}' already exists.", parent=self.master)
        else:
            messagebox.showwarning("Input Error", "Please enter a name for the new reviewer.", parent=self.master)


    def delete_reviewer(self):
        selected_indices = self.reviewer_listbox.curselection()
        if not selected_indices: messagebox.showwarning("Selection Error", "Please select a reviewer to delete.", parent=self.master); return
        selected_name = self.reviewer_listbox.get(selected_indices[0])
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete reviewer '{selected_name}'?", parent=self.master):
            if selected_name in self.all_reviewers: self.all_reviewers.remove(selected_name)
            self.refresh_reviewer_listbox()
            self.update_reviewer_checkboxes()
            self.update_status(f"Admin: Deleted '{selected_name}'. Click 'Save List' to make permanent.")


    def save_reviewer_list(self):
        if save_reviewers(self.all_reviewers):
            self.update_status("Admin: Reviewer list saved successfully.")
            messagebox.showinfo("Saved", f"Reviewer list saved to '{CONFIG_FILE}'.", parent=self.master)
        else:
             self.update_status("Admin: Error saving reviewer list.")


    def browse_lifesphere(self):
        filename = filedialog.askopenfilename(title="Select Lifesphere Export File", filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*")))
        if filename:
            self.lifesphere_file.set(filename)
            self.update_status(f"Selected Lifesphere file: {os.path.basename(filename)}")
            sheet_name = self.ls_sheet_name.get().strip()
            summary_msg = f"Selected Lifesphere file: {os.path.basename(filename)}\n"
            if not sheet_name: summary_msg += "ERROR: Please enter the sheet name."
            else:
                try:
                    df_temp = pd.read_excel(filename, sheet_name=sheet_name, skiprows=5, usecols=[0], dtype={0: str})
                    num_rows = df_temp.shape[0]
                    summary_msg += f"Found {num_rows} data rows in sheet '{sheet_name}' (after skipping 5 header rows)."
                except FileNotFoundError: summary_msg += f"ERROR: File not found."
                except ValueError as e: summary_msg += f"ERROR reading sheet '{sheet_name}'. Check name/format. ({e})"
                except Exception as e: summary_msg += f"ERROR reading '{sheet_name}'. ({e})"
            self.show_summary(summary_msg)


    def browse_previous(self):
        filename = filedialog.askopenfilename(title="Select Previous Assignment File", filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*")))
        if filename:
            self.prev_assign_file.set(filename)
            self.update_status(f"Selected Previous Assignment file: {os.path.basename(filename)}")
            prio_sheet = self.prev_prio_sheet_name.get().strip()
            pend_sheet = self.prev_pend_sheet_name.get().strip()
            summary_msg = f"Selected Previous Assignment file: {os.path.basename(filename)}\n"
            if not prio_sheet: summary_msg += "ERROR: Please enter the Priority sheet name.\n"
            else:
                try:
                    df_prio = pd.read_excel(filename, sheet_name=prio_sheet, usecols=[0])
                    summary_msg += f"Found {df_prio.shape[0]} rows in Priority sheet '{prio_sheet}'.\n"
                except ValueError as e: summary_msg += f"ERROR reading sheet '{prio_sheet}'. Check name/format. ({e})\n"
                except Exception as e: summary_msg += f"ERROR reading '{prio_sheet}'. ({e})\n"
            if not pend_sheet: summary_msg += "ERROR: Please enter the Pending sheet name.\n"
            else:
                 try:
                    df_pend = pd.read_excel(filename, sheet_name=pend_sheet, usecols=[0])
                    summary_msg += f"Found {df_pend.shape[0]} rows in Pending sheet '{pend_sheet}'."
                 except ValueError as e: summary_msg += f"ERROR reading sheet '{pend_sheet}'. Check name/format. ({e})"
                 except Exception as e: summary_msg += f"ERROR reading '{pend_sheet}'. ({e})"
            self.show_summary(summary_msg)


    def update_status(self, message):
        self.status_label.config(text=f"Status: {message}")
        self.master.update_idletasks()


    def show_summary(self, summary_message):
        self.summary_text.config(state=tk.NORMAL)
        self.summary_text.delete('1.0', tk.END)
        self.summary_text.insert(tk.END, summary_message)
        self.summary_text.config(state=tk.DISABLED)
        self.master.update_idletasks()


    def run_assignment_process(self):
        ls_file = self.lifesphere_file.get()
        prev_file = self.prev_assign_file.get()
        ls_sheet = self.ls_sheet_name.get().strip()
        prev_prio_sheet = self.prev_prio_sheet_name.get().strip()
        prev_pend_sheet = self.prev_pend_sheet_name.get().strip()

        # Validation
        if not ls_file or not os.path.exists(ls_file): messagebox.showerror("Input Error", "Please select a valid Lifesphere Export file.", parent=self.master); return
        if not ls_sheet: messagebox.showerror("Input Error", "Please enter the Lifesphere sheet name.", parent=self.master); return
        if prev_file:
            if not os.path.exists(prev_file): messagebox.showerror("Input Error", f"Previous Assignment file does not exist:\n{prev_file}", parent=self.master); return
            if not prev_prio_sheet: messagebox.showerror("Input Error", "Previous file selected, please enter the Priority sheet name.", parent=self.master); return
            if not prev_pend_sheet: messagebox.showerror("Input Error", "Previous file selected, please enter the Pending sheet name.", parent=self.master); return

        selected_reviewers = [name for name, var in self.reviewer_vars.items() if var.get()]
        if not selected_reviewers: messagebox.showerror("Input Error", "Please select at least one reviewer.", parent=self.master); return

        output_dir = filedialog.askdirectory(title="Select Directory to Save Output File")
        if not output_dir: self.update_status("Operation cancelled: No output directory selected."); return

        # Run Backend
        self.update_status("Processing... Please wait.")
        self.show_summary("Processing... Please wait.\nCheck console for detailed logs.")
        self.master.config(cursor="watch"); self.master.update()

        success, message = perform_assignment(
            ls_file, prev_file, ls_sheet, prev_prio_sheet, prev_pend_sheet,
            selected_reviewers, output_dir
        )

        self.master.config(cursor=""); self.update_status("Processing complete.")
        self.show_summary(message) # Display final summary/error

        if success: messagebox.showinfo("Success", "Assignment process completed successfully!", parent=self.master)
        else:
            formatted_message = '\n'.join(message[i:i+80] for i in range(0, len(message), 80)) # Basic wrap
            messagebox.showerror("Error", f"Assignment process failed:\n\n{formatted_message}", parent=self.master)


# --- Main Execution Block ---
if __name__ == "__main__":
    root = tk.Tk()
    try: # Optional theme setup
        style = ttk.Style(root)
        available_themes = style.theme_names()
        if 'vista' in available_themes: style.theme_use('vista')
        elif 'clam' in available_themes: style.theme_use('clam')
    except Exception as e: print(f"Could not apply custom theme: {e}")

    app = CaseAssignerApp(root)
    root.mainloop()